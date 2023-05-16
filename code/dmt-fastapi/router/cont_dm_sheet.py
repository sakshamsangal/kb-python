import time

import pandas as pd
from fastapi import APIRouter
# from fastapi import FastAPI
from pydantic import BaseModel

from openpyxl import load_workbook
from starlette.responses import StreamingResponse

import service.mapping as mapping
import service.filling as filling
import service.master as ms1
import service.xml_split as chunk

# from fastapi.responses import StreamingResponse
# from fastapi.staticfiles import StaticFiles

dm = APIRouter()


class XMLChunk(BaseModel):
    loc: str
    ct: str
    tag_selected: object
    all_dir: bool
    prod_names: list


class MapXpath(BaseModel):
    loc: str
    ct: str
    fn: str
    sn: str


class MSFileName(BaseModel):
    fn: str
    sn: str


class DMFileName(BaseModel):
    fn: str
    sn: str


class DMSheet(BaseModel):
    loc: str
    ct: str
    dm: DMFileName
    ms: MSFileName


class LocCt(BaseModel):
    loc: str
    ct: str


class CreateMaster(BaseModel):
    ct: str
    loc: str
    all_dir: str
    prod: list
    master_type: str


@dm.post("/dm-sheet", status_code=200)
async def dm_sheet(item: DMSheet):
    wb = load_workbook(f'{item.loc}/{item.ct}/excel/{item.ct}_rule.xlsx',
                       read_only=True)  # open an Excel file and return a workbook
    if 'tag_master' not in wb.sheetnames:
        ms1.tag_master(item.loc, item.ct, item.ms.fn, item.ms.sn)
    if 'tag_map' not in wb.sheetnames:
        df = pd.DataFrame(columns=['tag', 'map_tag'])
        df_tm = pd.read_excel(f'{item.loc}/{item.ct}/excel/{item.ct}_rule.xlsx', sheet_name='tag_master')
        df['tag'] = df_tm['tag']
        with pd.ExcelWriter(f'{item.loc}/{item.ct}/excel/{item.ct}_rule.xlsx', engine='openpyxl', mode='a',
                            if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name='tag_map', index=False)
        filling.fill_tm_by_dd(item.loc, item.ct)
    mapping.map_xpath(item.loc, item.ct, item.dm.fn, item.dm.sn)
    mapping.map_tag(item.loc, item.ct)
    filling.fill_feat(item.loc, item.ct)
    x = filling.fill_comp_style(item.loc, item.ct)
    return x


@dm.post("/tag-master", status_code=200)
async def tag_master(item: MapXpath):
    res = ms1.tag_master(item.loc, item.ct, item.fn, item.sn)
    return {'status': res}


@dm.post("/map-xpath", status_code=200)
async def map_xpath(item: MapXpath):
    res = mapping.map_xpath(item.loc, item.ct, item.fn, item.sn)
    return {'status': res}


@dm.post("/fill-tm-by-dd", status_code=200)
async def fill_tm_by_dd(item: LocCt):
    res = filling.fill_tm_by_dd(item.loc, item.ct)
    return {'status': res}


@dm.post("/map-tag", status_code=200)
async def map_tag(item: LocCt):
    res = mapping.map_tag(item.loc, item.ct)
    return {'status': res}


@dm.post("/create-dd", status_code=200)
async def create_dd(item: LocCt):
    res = mapping.create_dd(item.loc, item.ct)
    return {'status': res}


@dm.post("/fill-feat", status_code=200)
async def fill_feat(item: LocCt):
    res = filling.fill_feat(item.loc, item.ct)
    return {'status': res}


@dm.post("/fill-comp-style", status_code=200)
async def fill_comp_style(item: LocCt):
    x = filling.fill_comp_style(item.loc, item.ct)
    return x


@dm.post("/master", status_code=200)
def process_master(item: CreateMaster):
    res = ms1.process_master(item.loc, item.ct, item.all_dir, item.prod, item.master_type)
    return {'status': res}

# '''
# .\venv\Scripts\activate
# uvicorn main:app --reload
# '''

#
# async def fake_video_streamer():
#     for i in range(20):
#         yield b"some fake video bytes<br>"
#         time.sleep(1)
#
#     # yield b"AKSHU<br>"
#     # time.sleep(2)
#     # yield b"some fake video bytes<br>"
#     # time.sleep(2)
#     # yield b"AKSHU<br>"
#
#
# @dm.get("/lavi")
# async def main():
#     return StreamingResponse(fake_video_streamer(), media_type="text/html")
